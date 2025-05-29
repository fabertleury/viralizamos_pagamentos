import psycopg2
from datetime import datetime
import pytz
from decimal import Decimal
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import json

# Database connection parameters
DB_PARAMS = {
    'dbname': 'railway',
    'user': 'postgres',
    'password': 'zacEqGceWerpWpBZZqttjamDOCcdhRbO',
    'host': 'shinkansen.proxy.rlwy.net',
    'port': '29036'
}

def get_transactions_for_dates(dates):
    try:
        # Connect to the database
        conn = psycopg2.connect(**DB_PARAMS)
        cur = conn.cursor()
        
        transactions = []
        
        for date in dates:
            # Set start and end of the day
            start_date = datetime(date.year, date.month, date.day, 0, 0, 0, tzinfo=pytz.UTC)
            end_date = datetime(date.year, date.month, date.day, 23, 59, 59, tzinfo=pytz.UTC)
            
            # Query to get all approved transactions for the day with payment request details
            query = """
            SELECT 
                t.id,
                t.created_at,
                t.amount,
                t.status,
                t.method,
                t.provider,
                t.external_id,
                pr.customer_name,
                pr.customer_email,
                pr.customer_phone,
                pr.service_name,
                pr.profile_username,
                t.metadata
            FROM transactions t
            JOIN payment_requests pr ON t.payment_request_id = pr.id
            WHERE t.status = 'approved'
            AND t.created_at >= %s
            AND t.created_at <= %s
            ORDER BY t.created_at
            """
            
            cur.execute(query, (start_date, end_date))
            results = cur.fetchall()
            
            for row in results:
                # Parse metadata if it exists
                metadata = {}
                if row[12]:
                    try:
                        metadata = json.loads(row[12])  # Convert string to dict using json
                    except:
                        metadata = {}
                
                transaction = {
                    'Data': row[1].strftime('%d/%m/%Y %H:%M:%S'),
                    'ID Transação': row[0],
                    'Valor': float(row[2]),
                    'Status': row[3],
                    'Método': row[4],
                    'Provedor': row[5],
                    'ID Externo': row[6],
                    'Nome Cliente': row[7],
                    'Email Cliente': row[8],
                    'Telefone Cliente': row[9],
                    'Serviço': row[10],
                    'Perfil': row[11],
                    'Tipo Serviço': metadata.get('service_type', ''),
                    'Quantidade Total': metadata.get('total_quantity', ''),
                    'É Serviço de Seguidores': 'Sim' if metadata.get('is_followers_service') else 'Não'
                }
                transactions.append(transaction)
        
        return transactions
        
    except Exception as e:
        print(f"Erro ao conectar ao banco de dados: {str(e)}")
        return []
    finally:
        if 'cur' in locals():
            cur.close()
        if 'conn' in locals():
            conn.close()

def export_to_excel(transactions):
    if not transactions:
        print("Nenhuma transação encontrada para exportar.")
        return
    
    # Create DataFrame
    df = pd.DataFrame(transactions)
    
    # Create Excel writer
    filename = f'transacoes_aprovadas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    
    # Write DataFrame to Excel
    df.to_excel(writer, sheet_name='Transações', index=False)
    
    # Get workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Transações']
    
    # Format headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Format columns
    for col in range(1, len(df.columns) + 1):
        column_letter = get_column_letter(col)
        worksheet.column_dimensions[column_letter].width = 20
    
    # Format currency column
    for row in range(2, len(df) + 2):
        cell = worksheet.cell(row=row, column=3)  # Valor column
        cell.number_format = '"R$ "#,##0.00'
    
    # Save the file
    writer.close()
    print(f"\nArquivo Excel gerado com sucesso: {filename}")
    print(f"Total de transações exportadas: {len(transactions)}")

def main():
    # Define the specific dates
    dates = [
        datetime(2025, 5, 9, tzinfo=pytz.UTC),
        datetime(2025, 5, 14, tzinfo=pytz.UTC),
        datetime(2025, 5, 15, tzinfo=pytz.UTC),
        datetime(2025, 5, 17, tzinfo=pytz.UTC),
        datetime(2025, 5, 20, tzinfo=pytz.UTC),
        datetime(2025, 5, 23, tzinfo=pytz.UTC)
    ]
    
    print("\nExportando transações aprovadas para Excel...")
    transactions = get_transactions_for_dates(dates)
    export_to_excel(transactions)

if __name__ == "__main__":
    main() 